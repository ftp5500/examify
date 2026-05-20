import io
import uuid
import tempfile
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from django.db.models import Q
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken

from .models import Exam, Student, AnswerSheet, Question, School, UserProfile
from .serializers import (
    ExamSerializer, StudentSerializer, AnswerSheetSerializer,
    QuestionSerializer, SchoolSerializer,
    PrincipalRegistrationSerializer, CustomTokenObtainPairSerializer,
)
from .pdf_generator import generate_answer_sheet
from .omr_engine import process_answer_sheet, calculate_score


# ══════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════

def get_user_schools(user):
    """يرجع المدارس التي للمستخدم صلاحية عليها"""
    profile = getattr(user, 'profile', None)
    if profile and profile.role == 'principal':
        return user.managed_schools.all()
    return user.teaching_schools.all()


# ══════════════════════════════════════════════════
#  Auth
# ══════════════════════════════════════════════════

class CustomTokenObtainPairView(TokenObtainPairView):
    """تسجيل دخول يرجّع بيانات المستخدم مع التوكن"""
    serializer_class = CustomTokenObtainPairSerializer


class RegisterPrincipalView(APIView):
    """تسجيل مدير مدرسة جديد"""
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = PrincipalRegistrationSerializer(data=request.data)
        if serializer.is_valid():
            result = serializer.save()
            user   = result['user']
            school = result['school']
            refresh = RefreshToken.for_user(user)
            return Response({
                'message': 'تم التسجيل بنجاح',
                'user': {
                    'id':       user.id,
                    'username': user.username,
                    'full_name': user.profile.full_name,
                    'role':     'principal',
                },
                'school': {
                    'id':   school.id,
                    'name': school.name,
                },
                'access':  str(refresh.access_token),
                'refresh': str(refresh),
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ══════════════════════════════════════════════════
#  ViewSets
# ══════════════════════════════════════════════════

class ExamViewSet(viewsets.ModelViewSet):
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user    = self.request.user
        profile = getattr(user, 'profile', None)
        if profile and profile.role == 'principal':
            return Exam.objects.filter(school__in=user.managed_schools.all())
        return Exam.objects.filter(teacher=user)

    def perform_create(self, serializer):
        serializer.save(teacher=self.request.user)


class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        schools  = get_user_schools(self.request.user)
        queryset = Student.objects.filter(school__in=schools)

        grade   = self.request.query_params.get('grade')
        section = self.request.query_params.get('section')
        search  = self.request.query_params.get('search')

        if grade:
            queryset = queryset.filter(grade=grade)
        if section:
            queryset = queryset.filter(section=section)
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(student_id__icontains=search)
            )

        return queryset.order_by('grade', 'section', 'name')


class AnswerSheetViewSet(viewsets.ModelViewSet):
    serializer_class = AnswerSheetSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user    = self.request.user
        profile = getattr(user, 'profile', None)
        if profile and profile.role == 'principal':
            return AnswerSheet.objects.filter(
                exam__school__in=user.managed_schools.all()
            )
        return AnswerSheet.objects.filter(exam__teacher=user)


class SchoolViewSet(viewsets.ModelViewSet):
    serializer_class = SchoolSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return get_user_schools(self.request.user)

    def perform_create(self, serializer):
        serializer.save(principal=self.request.user)


class QuestionViewSet(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]


# ══════════════════════════════════════════════════
#  توليد أوراق PDF
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_sheets(request, exam_id):
    exam        = Exam.objects.get(id=exam_id)
    student_ids = request.data.get('student_ids', [])
    students    = Student.objects.filter(id__in=student_ids)

    if not students.exists():
        return Response({'error': 'no students selected'}, status=400)

    from PyPDF2 import PdfWriter, PdfReader

    writer = PdfWriter()

    for student in students:
        sheet, created = AnswerSheet.objects.get_or_create(
            exam=exam,
            student=student,
            defaults={'barcode': str(uuid.uuid4())[:12]}
        )
        buffer = generate_answer_sheet(exam, student, sheet.barcode)
        reader = PdfReader(buffer)
        writer.add_page(reader.pages[0])

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="sheets.pdf"'
    return response


# ══════════════════════════════════════════════════
#  تصحيح الورقة
# ══════════════════════════════════════════════════


"""
استبدل دالة grade_sheet الموجودة في views.py بهذه النسخة المحدّثة.
تحفظ صورة debug (warped + دوائر ملونة) في sheet.sheet_image كـ base64.
"""

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def grade_sheet(request, exam_id):
    import base64
    import cv2
    import numpy as np
    from .omr_engine import (
        correct_image, detect_bubbles, grade_by_circles,
        read_barcode, calculate_score
    )

    exam          = Exam.objects.get(id=exam_id)
    uploaded_file = request.FILES.get('image')

    if not uploaded_file:
        return Response({'error': 'no file'}, status=400)

    file_ext = os.path.splitext(uploaded_file.name)[1].lower()

    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
        for chunk in uploaded_file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        if file_ext == '.pdf':
            from pdf2image import convert_from_path
            pages    = convert_from_path(tmp_path, dpi=300)
            img_tmp  = tempfile.NamedTemporaryFile(delete=False, suffix='.jpg')
            pages[0].save(img_tmp.name, 'JPEG')
            process_path = img_tmp.name
        else:
            process_path = tmp_path

        # ── معالجة الصورة ──
        img = cv2.imread(process_path)
        if img is None:
            return Response({'error': 'فشل قراءة الصورة'}, status=400)

        if img.shape[1] > img.shape[0]:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)

        gray    = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        barcode = read_barcode(gray)
        warped  = correct_image(gray)

        clahe  = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        warped = clahe.apply(cv2.GaussianBlur(warped, (3, 3), 0))

        circles = detect_bubbles(warped, exam.num_choices)
        answers = {}

        if circles:
            answers = grade_by_circles(warped, circles, exam.num_questions, exam.num_choices)

        # ── بناء صورة debug ──
        debug_img = cv2.cvtColor(warped, cv2.COLOR_GRAY2BGR)
        if circles:
            h_w, w_w = warped.shape
            avg_r    = int(np.median([c[2] for c in circles]))
            mid_x    = w_w / 2

            # رسم كل الدوائر بالأحمر
            for cx, cy, r in circles:
                cv2.circle(debug_img, (cx, cy), r, (0, 0, 255), 1)

            # رسم الإجابة المختارة بالأخضر
            right_c = sorted([c for c in circles if c[0] > mid_x], key=lambda c: c[1])
            left_c  = sorted([c for c in circles if c[0] <= mid_x], key=lambda c: c[1])

            def cluster(lst):
                if not lst:
                    return []
                s   = sorted(lst, key=lambda c: c[1])
                gap = avg_r * 1.8
                rows, cur = [], [s[0]]
                for c in s[1:]:
                    if abs(c[1] - np.mean([r[1] for r in cur])) < gap:
                        cur.append(c)
                    else:
                        rows.append(cur)
                        cur = [c]
                rows.append(cur)
                return rows

            q_per_col   = (exam.num_questions + 1) // 2
            right_valid = sorted(
                [r for r in cluster(right_c) if abs(len(r) - exam.num_choices) <= 1],
                key=lambda r: np.mean([c[1] for c in r])
            )[:q_per_col]
            left_valid = sorted(
                [r for r in cluster(left_c) if abs(len(r) - exam.num_choices) <= 1],
                key=lambda r: np.mean([c[1] for c in r])
            )[:q_per_col]
            all_rows_dbg = (right_valid + left_valid)[:exam.num_questions]

            latin_keys = ['A', 'B', 'C', 'D', 'E']
            for idx, row in enumerate(all_rows_dbg):
                q_str = str(idx + 1)
                ans   = answers.get(q_str)
                if ans and ans in latin_keys:
                    row_sorted = sorted(row[:exam.num_choices], key=lambda c: -c[0])
                    j          = latin_keys.index(ans)
                    if j < len(row_sorted):
                        cx, cy, r = row_sorted[j]
                        cv2.circle(debug_img, (cx, cy), r + 4, (0, 255, 0), 3)

        # ── تحويل لـ base64 ──
        _, buf      = cv2.imencode('.jpg', debug_img, [cv2.IMWRITE_JPEG_QUALITY, 75])
        image_b64   = f"data:image/jpeg;base64,{base64.b64encode(buf).decode('utf-8')}"

        # ── قراءة الباركود ──
        if not barcode:
            barcode = request.data.get('barcode')

        if not barcode:
            return Response({'error': 'لم يتم قراءة الباركود', 'debug_image': image_b64}, status=400)

        sheet   = AnswerSheet.objects.get(barcode=barcode, exam=exam)
        correct = {str(q.number): q.correct_answer for q in exam.questions.all()}
        score   = calculate_score(answers, correct)

        sheet.answers     = answers
        sheet.score       = score
        sheet.status      = 'graded'
        sheet.sheet_image = image_b64
        sheet.save()

        return Response({
            'student':    sheet.student.name,
            'sheet_id':   sheet.id,
            'score':      score,
            'answers':    answers,
            'debug_image': image_b64,
        })
    finally:
        os.unlink(tmp_path)


# ══════════════════════════════════════════════════
#  تصدير النتائج Excel
# ══════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_results(request, exam_id):
    exam   = Exam.objects.get(id=exam_id)
    sheets = AnswerSheet.objects.filter(
        exam=exam, status='graded'
    ).select_related('student')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "النتائج"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells('A1:F1')
    ws['A1'] = f"نتائج {exam.title} - {exam.subject}"
    ws['A1'].font      = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['م', 'اسم الطالب', 'الرقم', 'الصف', 'الدرجة', 'التقدير']
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='2563EB')
        cell.alignment = Alignment(horizontal='center')

    for row_num, sheet in enumerate(sheets, 1):
        score = sheet.score or 0
        if score >= 90:
            grade = 'ممتاز'
        elif score >= 80:
            grade = 'جيد جداً'
        elif score >= 70:
            grade = 'جيد'
        elif score >= 60:
            grade = 'مقبول'
        else:
            grade = 'ضعيف'

        row = [row_num, sheet.student.name, sheet.student.student_id,
               sheet.student.grade, score, grade]
        for col, val in enumerate(row, 1):
            cell           = ws.cell(row=row_num + 2, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            if col == 5:
                cell.font = Font(bold=True,
                                 color='16A34A' if score >= 60 else 'DC2626')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="results_{exam_id}.xlsx"'
    return response


# ══════════════════════════════════════════════════
#  استيراد الطلاب من نور
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_students(request):
    excel_file = request.FILES.get('file')
    if not excel_file:
        return Response({'error': 'الملف مطلوب'}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        return Response({'error': f'ملف غير صالح: {str(e)}'}, status=400)

    # تحديد المدرسة
    school_id = request.data.get('school_id')
    if school_id:
        try:
            school = School.objects.get(id=school_id, principal=request.user)
        except School.DoesNotExist:
            return Response({'error': 'المدرسة غير موجودة'}, status=404)
    else:
        school = request.user.managed_schools.first()
        if not school:
            return Response({'error': 'لا توجد مدرسة مرتبطة بحسابك'}, status=400)

    total_created = 0
    total_updated = 0
    all_errors    = []

    def get_val(rows, row_idx, col_idx):
        try:
            v = rows[row_idx][col_idx]
            return str(v).strip() if v else ''
        except Exception:
            return ''

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 27:
            continue

        academic_year = get_val(rows, 1, 5)
        grade         = get_val(rows, 6, 3)
        section       = get_val(rows, 14, 2)
        school_name   = get_val(rows, 15, 37)

        if not school_name:
            continue

        student_rows = rows[26:]
        i = 0
        while i < len(student_rows):
            row = student_rows[i]
            seq = row[45] if len(row) > 45 else None
            if seq is None or str(seq).strip() == '':
                i += 1
                continue
            try:
                name       = str(row[40]).strip() if len(row) > 40 and row[40] else ''
                student_id = str(row[33]).strip() if len(row) > 33 and row[33] else ''

                if not name or not student_id:
                    i += 2
                    continue

                student, was_created = Student.objects.update_or_create(
                    student_id=student_id,
                    defaults={
                        'name':    name,
                        'grade':   grade,
                        'section': section,
                        'school':  school,
                    }
                )
                if was_created:
                    total_created += 1
                else:
                    total_updated += 1
            except Exception as e:
                all_errors.append(f"{sheet_name}: {str(e)}")
            i += 2

    return Response({
        'created': total_created,
        'updated': total_updated,
        'sheets':  len(wb.sheetnames),
        'errors':  all_errors[:5],
    })


# ══════════════════════════════════════════════════
#  استيراد المعلمين
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_teachers(request):
    """استيراد المعلمين من ملف Excel الرسمي"""
    from django.contrib.auth.models import User
    from django.core.mail import send_mail
    from django.conf import settings
    from .models import generate_random_password

    profile = getattr(request.user, 'profile', None)
    if not profile or profile.role != 'principal':
        return Response({'error': 'فقط مدير المدرسة يقدر يستورد معلمين'}, status=403)

    excel_file = request.FILES.get('file')
    school_id  = request.data.get('school_id')

    if not excel_file:
        return Response({'error': 'الملف مطلوب'}, status=400)
    if not school_id:
        return Response({'error': 'يجب اختيار المدرسة'}, status=400)

    try:
        school = School.objects.get(id=school_id, principal=request.user)
    except School.DoesNotExist:
        return Response({'error': 'المدرسة غير موجودة أو ليست لك صلاحية عليها'}, status=404)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        return Response({'error': f'ملف غير صالح: {str(e)}'}, status=400)

    ws      = wb.active
    created = []
    linked  = []
    errors  = []
    email_failures = []

    def normalize_mobile(mobile):
        if not mobile:
            return ''
        m = str(mobile).strip().replace(' ', '').replace('-', '')
        if m.startswith('00966'):
            return '+966' + m[5:]
        if m.startswith('966'):
            return '+' + m
        if m.startswith('05'):
            return '+966' + m[1:]
        if m.startswith('5') and len(m) == 9:
            return '+966' + m
        return m

    for row_idx in range(16, ws.max_row + 1):
        name        = ws.cell(row=row_idx, column=22).value
        national_id = ws.cell(row=row_idx, column=24).value
        mobile      = ws.cell(row=row_idx, column=6).value
        email       = ws.cell(row=row_idx, column=9).value

        if not name or not national_id:
            continue

        name        = str(name).strip()
        national_id = str(national_id).strip()
        mobile      = normalize_mobile(mobile)
        email       = str(email).strip() if email else ''

        if not national_id.isdigit() or len(national_id) != 10:
            errors.append(f"رقم هوية غير صالح: {name} ({national_id})")
            continue

        try:
            existing_profile = UserProfile.objects.filter(national_id=national_id).first()

            if existing_profile:
                if existing_profile.user not in school.teachers.all():
                    school.teachers.add(existing_profile.user)
                    linked.append({'name': name, 'national_id': national_id})
            else:
                password = generate_random_password()

                user = User.objects.create_user(
                    username=national_id,
                    email=email or '',
                    password=password,
                    first_name=name.split()[0] if name else '',
                )

                UserProfile.objects.create(
                    user=user,
                    role='teacher',
                    full_name=name,
                    national_id=national_id,
                    mobile=mobile,
                    must_change_password=True,
                )

                school.teachers.add(user)

                if email:
                    try:
                        send_mail(
                            subject=f'تم تسجيلك في نظام Examify - {school.name}',
                            message=(
                                f'مرحباً أ. {name}،\n\n'
                                f'تم تسجيلك كمعلم في {school.name}.\n\n'
                                f'بيانات الدخول:\n'
                                f'اسم المستخدم: {national_id}\n'
                                f'كلمة المرور المؤقتة: {password}\n\n'
                                f'⚠️ سيُطلب منك تغيير كلمة المرور عند أول تسجيل دخول.\n\n'
                                f'رابط النظام: http://localhost:3000\n\n'
                                f'مع تحيات إدارة {school.name}'
                            ),
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            recipient_list=[email],
                            fail_silently=False,
                        )
                    except Exception as e:
                        email_failures.append({'name': name, 'email': email, 'error': str(e)})

                created.append({
                    'name':         name,
                    'national_id':  national_id,
                    'email':        email,
                    'temp_password': password if not email else None,
                })

        except Exception as e:
            errors.append(f"{name}: {str(e)}")

    return Response({
        'created_count':       len(created),
        'linked_count':        len(linked),
        'errors_count':        len(errors),
        'email_failures_count': len(email_failures),
        'created':             created,
        'linked':              linked,
        'errors':              errors[:10],
        'email_failures':      email_failures[:10],
    })


# ══════════════════════════════════════════════════
#  قائمة المعلمين
# ══════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_teachers(request):
    """قائمة معلمي مدارس المدير"""
    from django.contrib.auth.models import User

    profile = getattr(request.user, 'profile', None)
    if not profile or (profile.role != 'principal' and not request.user.is_superuser):
        return Response({'error': 'فقط المدير يقدر يشوف المعلمين'}, status=403)

    schools  = request.user.managed_schools.all()
    teachers = User.objects.filter(teaching_schools__in=schools).distinct()

    data = []
    for t in teachers:
        p = getattr(t, 'profile', None)
        data.append({
            'id':                   t.id,
            'username':             t.username,
            'email':                t.email,
            'full_name':            p.full_name if p else '',
            'national_id':          p.national_id if p else '',
            'mobile':               p.mobile if p else '',
            'must_change_password': p.must_change_password if p else False,
        })
    return Response(data)


# ══════════════════════════════════════════════════
#  إعادة إرسال بيانات دخول معلم
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def resend_teacher_credentials(request, teacher_id):
    """إرسال أو إعادة إرسال بيانات الدخول لمعلم"""
    from django.contrib.auth.models import User
    from django.core.mail import send_mail
    from django.conf import settings
    from .models import generate_random_password

    profile = getattr(request.user, 'profile', None)
    if not profile or profile.role != 'principal':
        return Response({'error': 'فقط المدير يقدر يرسل بيانات الدخول'}, status=403)

    try:
        teacher = User.objects.get(id=teacher_id)
    except User.DoesNotExist:
        return Response({'error': 'المعلم غير موجود'}, status=404)

    schools = request.user.managed_schools.all()
    if not teacher.teaching_schools.filter(id__in=schools).exists():
        return Response({'error': 'هذا المعلم ليس في مدرستك'}, status=403)

    teacher_profile = getattr(teacher, 'profile', None)
    if not teacher_profile:
        return Response({'error': 'بيانات المعلم غير كاملة'}, status=400)

    if not teacher.email:
        return Response({'error': 'المعلم ليس لديه بريد إلكتروني مسجل'}, status=400)

    new_password = generate_random_password()
    teacher.set_password(new_password)
    teacher.save()

    teacher_profile.must_change_password = True
    teacher_profile.save()

    school = schools.first()
    try:
        send_mail(
            subject=f'بيانات الدخول إلى نظام Examify - {school.name}',
            message=(
                f'مرحباً أ. {teacher_profile.full_name}،\n\n'
                f'تم إعادة تعيين كلمة المرور الخاصة بك في نظام Examify.\n\n'
                f'بيانات الدخول:\n'
                f'اسم المستخدم: {teacher.username}\n'
                f'كلمة المرور الجديدة: {new_password}\n\n'
                f'⚠️ سيُطلب منك تغيير كلمة المرور عند تسجيل الدخول.\n\n'
                f'رابط النظام: http://localhost:3000\n\n'
                f'مع تحيات إدارة {school.name}'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[teacher.email],
            fail_silently=False,
        )
        return Response({
            'success':      True,
            'message':      f'تم إرسال بيانات الدخول إلى {teacher.email}',
            'temp_password': new_password,
        })
    except Exception as e:
        return Response({
            'error':        f'فشل إرسال البريد الإلكتروني: {str(e)}',
            'temp_password': new_password,
        }, status=500)


# ══════════════════════════════════════════════════
#  تغيير كلمة السر
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    """تغيير كلمة السر — إجباري لأول دخول"""
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')

    if not old_password or not new_password:
        return Response({'error': 'كلمة المرور القديمة والجديدة مطلوبتان'}, status=400)

    if not request.user.check_password(old_password):
        return Response({'error': 'كلمة المرور القديمة غير صحيحة'}, status=400)

    if len(new_password) < 8:
        return Response({'error': 'كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل'}, status=400)

    if old_password == new_password:
        return Response({'error': 'كلمة المرور الجديدة يجب أن تختلف عن القديمة'}, status=400)

    request.user.set_password(new_password)
    request.user.save()

    profile = getattr(request.user, 'profile', None)
    if profile:
        profile.must_change_password = False
        profile.save()

    return Response({'message': 'تم تغيير كلمة المرور بنجاح'})


# ══════════════════════════════════════════════════
#  فلاتر الطلاب
# ══════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_filters(request):
    """يرجع الصفوف والفصول المتاحة"""
    schools  = get_user_schools(request.user)
    queryset = Student.objects.filter(school__in=schools)
    grades   = list(queryset.values_list('grade', flat=True).distinct().order_by('grade'))
    return Response({'grades': [g for g in grades if g]})


# ══════════════════════════════════════════════════
#  تفاصيل طالب
# ══════════════════════════════════════════════════

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def student_details(request, student_id):
    """تفاصيل طالب مع سجل اختباراته كاملاً"""
    schools = get_user_schools(request.user)

    try:
        student = Student.objects.get(id=student_id, school__in=schools)
    except Student.DoesNotExist:
        return Response({'error': 'الطالب غير موجود'}, status=404)

    sheets = AnswerSheet.objects.filter(
        student=student
    ).select_related('exam').order_by('-id')

    history = []
    for sheet in sheets:
        history.append({
            'sheet_id':    sheet.id,
            'exam_id':     sheet.exam.id,
            'exam_title':  sheet.exam.title,
            'exam_subject': sheet.exam.subject,
            'exam_date':   str(sheet.exam.exam_date) if sheet.exam.exam_date else None,
            'status':      sheet.status,
            'score':       sheet.score,
        })

    graded    = [s for s in history if s['status'] == 'graded' and s['score'] is not None]
    avg_score = sum(s['score'] for s in graded) / len(graded) if graded else None

    return Response({
        'id':         student.id,
        'name':       student.name,
        'student_id': student.student_id,
        'grade':      student.grade,
        'section':    student.section,
        'school':     student.school.name if student.school else '',
        'stats': {
            'total':     len(history),
            'graded':    len(graded),
            'pending':   len([s for s in history if s['status'] == 'pending']),
            'avg_score': round(avg_score, 1) if avg_score is not None else None,
        },
        'history': history,
    })


# ══════════════════════════════════════════════════
#  تعديل إجابات ورقة طالب
# ══════════════════════════════════════════════════

@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def update_sheet_answers(request, sheet_id):
    """تعديل إجابات الطالب وإعادة حساب الدرجة تلقائياً"""
    try:
        sheet = AnswerSheet.objects.get(id=sheet_id)
    except AnswerSheet.DoesNotExist:
        return Response({'error': 'الورقة غير موجودة'}, status=404)

    new_answers = request.data.get('answers')
    if not new_answers:
        return Response({'error': 'الإجابات مطلوبة'}, status=400)

    sheet.answers = new_answers
    correct       = {str(q.number): q.correct_answer for q in sheet.exam.questions.all()}
    sheet.score   = calculate_score(new_answers, correct)
    sheet.status  = 'graded'
    sheet.save()

    return Response({
        'answers': sheet.answers,
        'score':   round(sheet.score, 1),
    })


# ══════════════════════════════════════════════════
#  Debug: تشخيص التصحيح
# ══════════════════════════════════════════════════

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def debug_grade(request, exam_id):
    """
    يصحح الورقة ويرجع صورة debug تظهر ما يراه النظام
    """
    import base64
    import cv2
    import numpy as np
    from .omr_engine import correct_image, detect_bubbles, grade_by_circles, read_barcode

    exam          = Exam.objects.get(id=exam_id)
    uploaded_file = request.FILES.get('image')
    if not uploaded_file:
        return Response({'error': 'لا يوجد ملف'}, status=400)

    file_ext = os.path.splitext(uploaded_file.name)[1].lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
        for chunk in uploaded_file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    try:
        img = cv2.imread(tmp_path)
        if img.shape[1] > img.shape[0]:
            img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        gray    = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        barcode = read_barcode(gray)
        warped  = correct_image(gray)
        clahe   = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        warped  = clahe.apply(cv2.GaussianBlur(warped, (3, 3), 0))

        circles = detect_bubbles(warped, exam.num_choices)

        answers    = {}
        debug_info = {'circles_found': len(circles) if circles else 0}

        if circles:
            answers = grade_by_circles(warped, circles, exam.num_questions, exam.num_choices)

            # ── نفس منطق grade_by_circles للعرض ──
            h_w, w_w = warped.shape
            avg_r    = int(np.median([c[2] for c in circles]))
            mid_x    = w_w / 2

            right_c = [c for c in circles if c[0] > mid_x]
            left_c  = [c for c in circles if c[0] <= mid_x]

            def cluster(lst):
                if not lst:
                    return []
                s   = sorted(lst, key=lambda c: c[1])
                gap = avg_r * 1.8
                rows, cur = [], [s[0]]
                for c in s[1:]:
                    if abs(c[1] - np.mean([r[1] for r in cur])) < gap:
                        cur.append(c)
                    else:
                        rows.append(cur)
                        cur = [c]
                rows.append(cur)
                return rows

            rr = cluster(right_c)
            lr = cluster(left_c)
            rv = [r for r in rr if abs(len(r) - exam.num_choices) <= 1]
            lv = [r for r in lr if abs(len(r) - exam.num_choices) <= 1]

            debug_info = {
                'circles_found':  len(circles),
                'right_circles':  len(right_c),
                'left_circles':   len(left_c),
                'right_rows':     len(rr),
                'left_rows':      len(lr),
                'right_valid':    len(rv),
                'left_valid':     len(lv),
                'right_lengths':  [len(r) for r in rr],
                'left_lengths':   [len(r) for r in lr],
                'avg_radius':     avg_r,
                'answers':        answers,
                'barcode':        barcode,
            }

            # ── readings لكل سؤال ──
            q_per_col2  = (exam.num_questions + 1) // 2
            right_sorted = sorted(rv, key=lambda r: np.mean([c[1] for c in r]))[:q_per_col2]
            left_sorted  = sorted(lv, key=lambda r: np.mean([c[1] for c in r]))[:q_per_col2]
            all_rows_dbg = (right_sorted + left_sorted)[:exam.num_questions]

            debug_readings = {}
            for idx, row in enumerate(all_rows_dbg):
                q_num      = idx + 1
                row_sorted = sorted(row[:exam.num_choices], key=lambda c: -c[0])
                readings   = []
                for cx, cy, r in row_sorted:
                    roi_r = int(r * 1.2)
                    roi   = warped[max(0, cy - roi_r):min(h_w, cy + roi_r),
                                   max(0, cx - roi_r):min(w_w, cx + roi_r)]
                    readings.append(round(float(np.mean(roi)) if roi.size > 0 else 255.0, 1))
                debug_readings[str(q_num)] = {
                    'readings': readings,
                    'min':      round(min(readings), 1),
                    'spread':   round(max(readings) - min(readings), 1),
                    'detected': answers.get(str(q_num)),
                }
            debug_info['readings'] = debug_readings

            # ── رسم debug ──
            debug_img = cv2.cvtColor(warped, cv2.COLOR_GRAY2BGR)
            for cx, cy, r in circles:
                cv2.circle(debug_img, (cx, cy), r, (0, 0, 255), 1)

            for idx, row in enumerate(
                sorted(rv, key=lambda r: np.mean([c[1] for c in r]))[:q_per_col2]
            ):
                for cx, cy, r in sorted(row, key=lambda c: -c[0])[:exam.num_choices]:
                    roi_r = int(r * 1.2)
                    roi   = warped[max(0, cy - roi_r):min(h_w, cy + roi_r),
                                   max(0, cx - roi_r):min(w_w, cx + roi_r)]
                    if roi.size > 0 and np.mean(roi) < 150:
                        cv2.circle(debug_img, (cx, cy), r + 3, (0, 255, 0), 2)

        else:
            debug_img = cv2.cvtColor(warped, cv2.COLOR_GRAY2BGR)

        _, buf  = cv2.imencode('.jpg', debug_img, [cv2.IMWRITE_JPEG_QUALITY, 80])
        img_b64 = base64.b64encode(buf).decode('utf-8')

        return Response({
            'debug_image': f'data:image/jpeg;base64,{img_b64}',
            'warped_size': f'{warped.shape[1]}×{warped.shape[0]}',
            'debug_info':  debug_info,
        })
    finally:
        os.unlink(tmp_path)